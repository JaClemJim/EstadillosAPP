import base64
import streamlit as st
import pandas as pd
from shift_scheduling_sat_revCREF_v20 import solve_shift_scheduling
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime, timedelta
import warnings
import myInputCRs

#Debido a que hay conflictos de compatibilidad entre versiones de protobuf, ortools y streamlit, aparecen warnings avisando que
#se instale la ultima versión de las mismas. Se evita con esta librería
warnings.filterwarnings("ignore")

st.set_page_config(layout="wide")

######
#Entradilla
######
st.write(""" # ESTADILLOS INTERACTIVOS """)

st.markdown("""
 Seleccione los parámetros:

 (Imprescindible rellenar nombre de aeropuerto primero) 
""")

######
# Carga de los resultados en caché. Acelera el funcionamiento de la página.
# se puede considerar st.cache_data según la versión de streamlit. 
# Entrada:
#   - Lista con los parámetros que se introducen en la web manualmente
# Salida:
#   - Dataframe resumen
#   - Dataframe con el estadillo
######
@st.cache
def load_data(datos, traf = []):
    lista = solve_shift_scheduling(datos, traf)
    return lista

@st.cache
def load_turnos(datos, ad, t_id):
    h=pd.read_csv(datos,sep=";")

    h_2 =  h.loc[h['ICAO']==ad,['turnoshini']]

    h_3 = [i.split(',') for i in h_2['turnoshini']]

    inicio = float(h_3[0][t_id])

    return inicio

def decimal_to_time(decimal_hour):
    hours = int(decimal_hour)
    minutes = int((decimal_hour - hours) * 60)
    time = datetime(1, 1, 1, hours, minutes)
    return time.strftime("%H:%M")

def generar_lista_hora(hora_inicial, num_columnas):
    hora_inicio = decimal_to_time(hora_inicial)
    hora_actual = datetime.strptime(hora_inicio, '%H:%M')
    lista_hora = ["H Inicio " + hora_inicio]

    for _ in range(num_columnas - 1):
        hora_actual += timedelta(minutes=bloque)
        lista_hora.append("H Inicio " + hora_actual.strftime('%H:%M'))

    return lista_hora

#####
# Transformar dataframe en excel descargable. 
# Entrada:
#   - tabla: dataframe con el estadillo
# salida:
#   - hoja de excel con el estadillo con el formato seleccionado
#####
@st.cache
def transf(tabla, tabla2, aeropuerto):
    wb = Workbook()
    ws1 = wb.active
    for r in dataframe_to_rows(tabla, index=False, header=False):
        ws1.append(r)
        

    # gradiente de colores en primera fila de estadillo
    for col_num, column_title in enumerate(tabla.columns[3:], 1):
        cell = ws1.cell(row=1, column=col_num)
        min_color = '25D82B' # Lightest color
        mid_color = 'F0A22A'
        max_color = 'E03C18' # Darkest color
        rule = ColorScaleRule(start_type='min', start_color=min_color,
                            mid_type='num', mid_value=70, mid_color=mid_color,
                            end_type='max', end_color=max_color)
        ws1.conditional_formatting.add('D2:CO2', rule)
        
    # Agrupa las celdas de número en la primera fila
    grupo = int(time/t_bloque)
    column_index = 4
    num_groups = (tabla.shape[1] - 3) // grupo
    last_group_size = (tabla.shape[1] - 3) % grupo

    for i in range(num_groups+1):
        if i == num_groups and last_group_size != 0:
            group_size = last_group_size
        else:
            group_size = grupo
            
        column_letter_start = get_column_letter(column_index)
        column_letter_end = get_column_letter(column_index + group_size - 1)
        cell_start_1 = f'{column_letter_start}1'
        cell_end_1 = f'{column_letter_end}1'  
        
        
        ws1.merge_cells(f'{cell_start_1}:{cell_end_1}')
        cell_start_2 = f'{column_letter_start}2'
        cell_end_2 = f'{column_letter_end}2'  
        ws1.merge_cells(f'{cell_start_2}:{cell_end_2}')    
        

        
        column_index += group_size
        
    # rellenar de verde las celdas en las que se trabaja
    for i in range(4, ws1.max_column + 1):
        for j in range(1, ws1.max_row + 1):
            cell = ws1.cell(row=j, column=i)         
            
            if cell.value == 'T':
                fill_color = PatternFill(start_color='91E183', end_color='91E183', fill_type='solid')
                cell.fill = fill_color

    # cambiar tamaño de columnas
    ws1.column_dimensions['A'].width = 20
    for i in range(4, 94):
        col_letter = get_column_letter(i)
        ws1.column_dimensions[col_letter].width = 2.8
    
    ws2 = wb.create_sheet()
    for t in dataframe_to_rows(tabla2, index=False, header=False):
        ws2.append(t)
        

    #generar excel descargable     
    wb.save(aeropuerto + ".xlsx") 
    
######
# Input
######

aerop = st.text_input("código OACI")
atcos = st.number_input('Número de ATCOS disponibles para el turno', min_value=1, step=1)
turno = st.selectbox('0 -> Mañana, 1 -> tarde, 2 -> noche', [0,1,2])
bloque = st.number_input('Bloque de tiempo para dividir la hora', min_value=5, max_value= 60, step=5)
demanda = st.number_input('Bloque de tiempo para captar la demanda', min_value=5, max_value = 60, step=5)
dia = st.date_input("fecha (por defecto, hoy)")


list_input = [aerop, atcos, turno, bloque, demanda, dia] #lista que sirve para alimentar a la fucnión que genera los estadillos
# st.write(list_input)

######
# demanda por hora para lista manipulable
######
check1 = st.checkbox("Selecciona el recuadro si quieres modificar la demanda a mano")
if check1:

    a = myInputCRs.MyEscenario(icao=aerop)
    list_demanda = a.getdfTrafico(diames=dia.day, idturno=turno, ventanaflotante=60) #demanda por hora
    new_list_demanda = []

    for i, num_demand in enumerate(list_demanda[0]):
        key = f"numero_{i}"
        nuevo_valor = st.number_input(label=f"Demanda en hora {i+1}", value=int(num_demand), key=key, step = 1)
        
        new_list_demanda.append(nuevo_valor)

    
    # st.write(new_list_demanda)



#######
# Ejecución del código
#######

boton1 = st.button("Click para calcular")

# st.write("boton:", boton1)

#cada vez que se hace click se ejecuta, sino no, si se cambia algún campo se reinicia y el código vuelve a esta línea
if boton1:
    # try:
    if check1:
        sol = load_data(list_input, traf = new_list_demanda)
    else:
        sol = load_data(list_input)


    if type(sol) == list:
        resultado = sol[0]
        mensaje = sol[1]

        if len(mensaje) !=0:
            st.write(mensaje[0])

        time = demanda #minutos
        t_bloque  = bloque #minutos

        # Formato de la salida de la función que calcula el estadillo
        dfs = [pd.DataFrame(line.split(',')).transpose() for line in resultado]
        df = pd.concat(dfs).reset_index(drop=True).iloc[:, 0:-1]

        grupo = int(time/t_bloque)

        lista1 = [i for i in list(filter(lambda x: x != ' ', df.iloc[0,1:])) for j in range(grupo)]
        lista2 = [i for i in list(filter(lambda x: x != ' ', df.iloc[1,1:])) for j in range(grupo)]

        if len(lista1) > (df.shape[1] - 1):
            a = len(lista1) - (df.shape[1] - 1)
            lista1 = lista1[:-a]
            lista2 = lista2[:-a]

        df.loc[0, 1:] = lista1
        df.loc[1, 1:] = lista2 

        df.loc[:1, 1:]=df.loc[:1, 1:].astype('int')

        durations = []
        for index, row in df.iterrows():
            duration = 0
            for value in row.values:
                if value == 'T':
                    duration += 1
            durations.append(duration)

        porcentaje = [(i/(len(df.columns)-1))*100 for i in durations]
        duration = [(i*t_bloque)/60 for i in durations]

        df.insert(1, 'tiempo', duration)
        df.insert(2, 'porcentaje', porcentaje)

        # print(df)

        df2 = df.iloc[2:,3:]
        count_dicc = {}

        for index, row in df2.iterrows():
            count_list = []
            current_item = row.values[0]
            current_count = t_bloque
            
            for item in row.values[1:]:
                if item == current_item:
                    current_count += t_bloque
                else:
                    last_item = current_item
                    count_list.append((last_item+':', current_count))
                    current_item = item
                    current_count = t_bloque
            
            count_list.append(((item+':', current_count)))
            
            count_dicc['worker'+str(index-2)] = count_list
            
        longitud_maxima = max(map(len, count_dicc.values()))

        for key in count_dicc:
            lista = count_dicc[key]
            while len(lista) < longitud_maxima:
                lista.append(None)

        df3 = pd.DataFrame(count_dicc).transpose().reset_index().replace({None: ''}).astype('str')

        # st.write(df3)

        df_copy = df.copy()

        titulos = list(df_copy.columns)

        h_ini = load_turnos("datosDependencias1.csv", aerop, turno)
        lista_horas = generar_lista_hora(h_ini, len(titulos))

        new_columns = ["ATCOS"] + titulos[1:3] + lista_horas


        # new_columns = []
        # for col in df_copy.columns:
        #     if isinstance(col, int) and col != 0:
        #         new_columns.append("Bloque nº " + str(col))
        #     elif col == 0:
        #         new_columns.append("ATCO")
        #     else:
        #         new_columns.append(col)
        
        df_copy.columns = new_columns

        for col in df_copy.columns:
            if "H inicio" in col:
                df_copy[col] = df_copy[col].astype(str)

        
        # # Definir una función para aplicar estilos personalizados a las celdas
        # def highlight_cells(value):
        #     style = 'background-color: green; color: white' if value == 'T' else ''
        #     return style

        # # Aplicar estilos personalizados al DataFrame
        # styled_df = df.style.apply(lambda row: highlight_cells(row), axis=1)

        # # Mostrar el DataFrame en Streamlit
        # st.dataframe(styled_df)

        # Definir una función para aplicar estilos personalizados a las celdas
        def highlight_cells(value):
            style = ['background-color: green; color: white' if v == 'T' else '' for v in value]
            return style

        # Aplicar estilos personalizados al DataFrame
        styled_df = df_copy.style.apply(highlight_cells, axis=0)

        # Mostrar el DataFrame en Streamlit
        st.dataframe(styled_df)

        #generar excel
        transf(df, df3, aerop)

        nombre = aerop+'.xlsx'

        # Abrir excel, codificar y generar enlace de descarga
        with open(nombre, 'rb') as f:
            estadillo = f.read()

        b64 = base64.b64encode(estadillo).decode()

        href = f'<a href="data:application/estadillo;base64,{b64}" download="{nombre}">Descargar Excel</a>'
        st.markdown(href, unsafe_allow_html=True)
    else:
        st.error(sol)

    # except Exception as e:
    #     # st.write(str(e))
    #     logging.error(str(e))
    #     st.error(f"{str(e)}")



